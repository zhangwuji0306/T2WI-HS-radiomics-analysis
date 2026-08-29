#!/usr/bin/env python3
"""Create a local, stable mapping from original image IDs to anonymous IDs.

The mapping is intentionally written below ``local_private/`` and this script
does not modify source images or source tables.
"""

from __future__ import annotations

import argparse
import csv
import re
import secrets
from pathlib import Path
from typing import Dict, List, Optional, Set


ID_HEADERS = {
    "影像号",
    "image_id",
    "image_no",
    "image_number",
    "case_id",
    "patient_id",
}
ORIGINAL_COLUMN = "original_image_id"
ANONYMOUS_COLUMN = "anonymous_image_id"
ID_PATTERN = re.compile(r"^\d{4,}$")


def normalize_id(value: object) -> Optional[str]:
    text = str(value).strip()
    if text.endswith(".0"):
        text = text[:-2]
    return text if ID_PATTERN.fullmatch(text) else None


def normalized_headers(fieldnames: Optional[List[Optional[str]]]) -> Dict[str, str]:
    result: Dict[str, str] = {}
    for field in fieldnames or []:
        if field is None:
            continue
        clean = field.strip().lstrip("\ufeff")
        result[clean.lower()] = field
    return result


def collect_from_csv(path: Path, ids: Set[str]) -> None:
    try:
        with path.open("r", encoding="utf-8-sig", errors="replace", newline="") as handle:
            reader = csv.DictReader(handle)
            headers = normalized_headers(reader.fieldnames)
            selected = [original for lower, original in headers.items() if lower in {h.lower() for h in ID_HEADERS}]
            for row in reader:
                for header in selected:
                    image_id = normalize_id(row.get(header, ""))
                    if image_id:
                        ids.add(image_id)
    except (OSError, UnicodeError, csv.Error):
        return


def collect_from_xlsx(path: Path, ids: Set[str]) -> None:
    try:
        from openpyxl import load_workbook
    except ImportError:
        return

    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
        for sheet in workbook.worksheets:
            rows = sheet.iter_rows(values_only=True)
            header_row = next(rows, None)
            if not header_row:
                continue
            headers = [str(value).strip().lstrip("\ufeff") if value is not None else "" for value in header_row]
            selected = [index for index, header in enumerate(headers) if header.lower() in {h.lower() for h in ID_HEADERS}]
            for row in rows:
                for index in selected:
                    if index < len(row):
                        image_id = normalize_id(row[index])
                        if image_id:
                            ids.add(image_id)
        workbook.close()
    except (OSError, ValueError, TypeError):
        return


def collect_ids(root: Path) -> Set[str]:
    ids: Set[str] = set()
    for relative in (Path("feature_extract") / "result", Path("feature_extract") / "reader_2_result"):
        directory = root / relative
        if directory.is_dir():
            for child in directory.iterdir():
                if child.is_dir():
                    image_id = normalize_id(child.name)
                    if image_id:
                        ids.add(image_id)

    source_roots = [root / name for name in ("feature_extract", "habitat_analysis", "prognosis_analysis", "archive")]
    for source_root in source_roots:
        if not source_root.is_dir():
            continue
        for path in source_root.rglob("*.csv"):
            collect_from_csv(path, ids)
        for path in source_root.rglob("*.xlsx"):
            collect_from_xlsx(path, ids)
    return ids


def load_existing(path: Path) -> Dict[str, str]:
    if not path.is_file():
        return {}
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        return {
            row[ORIGINAL_COLUMN]: row[ANONYMOUS_COLUMN]
            for row in reader
            if row.get(ORIGINAL_COLUMN) and row.get(ANONYMOUS_COLUMN)
        }


def create_mapping(root: Path, output: Path) -> None:
    mapping = load_existing(output)
    used = set(mapping.values())
    for image_id in sorted(collect_ids(root)):
        if image_id in mapping:
            continue
        while True:
            anonymous_id = "IMG-" + secrets.token_hex(5).upper()
            if anonymous_id not in used:
                break
        mapping[image_id] = anonymous_id
        used.add(anonymous_id)

    output.parent.mkdir(parents=True, exist_ok=True)
    with output.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=[ORIGINAL_COLUMN, ANONYMOUS_COLUMN])
        writer.writeheader()
        for image_id in sorted(mapping):
            writer.writerow({ORIGINAL_COLUMN: image_id, ANONYMOUS_COLUMN: mapping[image_id]})
    print("Wrote %d mappings to %s" % (len(mapping), output))


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--root", type=Path, default=Path(__file__).resolve().parents[1])
    parser.add_argument("--output", type=Path, default=None)
    args = parser.parse_args()
    output = args.output or args.root / "local_private" / "image_id_mapping.csv"
    create_mapping(args.root.resolve(), output.resolve())


if __name__ == "__main__":
    main()
