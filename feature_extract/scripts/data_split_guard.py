"""Central cohort split and fail-closed data-access helpers.

The three public readers deliberately separate technical A, A outcome, and B
validation access.  Default CSV/XLSX reads are row-streamed and filtered by an
already-authorized identifier allow-list before a pandas frame is created.
Arbitrary injected readers are rejected because a callable returning a mixed
frame cannot prove that disallowed rows were not materialized first.
"""
from __future__ import annotations

import csv
import hashlib
import math
import os
import re
import sys
from typing import Iterable, Optional, Set

import pandas as pd

try:
    from openpyxl.utils.cell import coordinate_to_tuple
    from openpyxl.worksheet._reader import WorkSheetParser
except ImportError:  # pragma: no cover - CSV readers remain importable without openpyxl
    coordinate_to_tuple = None
    WorkSheetParser = None

HERE = os.path.dirname(os.path.abspath(__file__))
FEATURE_ROOT = os.path.dirname(HERE)
PROJECT_ROOT = os.path.dirname(FEATURE_ROOT)
HABITAT_ROOT = os.path.join(PROJECT_ROOT, "habitat_analysis")
HABITAT_SCRIPTS = os.path.join(HABITAT_ROOT, "scripts")
if HABITAT_SCRIPTS not in sys.path:
    sys.path.insert(0, HABITAT_SCRIPTS)
from freeze_lock import validate_freeze_lock  # noqa: E402

PROGNOSIS_ROOT = os.path.join(PROJECT_ROOT, "prognosis_analysis")
PROGNOSIS_SCRIPTS = os.path.join(PROGNOSIS_ROOT, "scripts")
if PROGNOSIS_SCRIPTS not in sys.path:
    sys.path.insert(0, PROGNOSIS_SCRIPTS)
from model_freeze_lock import validate_model_freeze_lock  # noqa: E402

FREEZE_LOCK = os.path.join(HABITAT_ROOT, "freeze_lock.json")
# Kept as a compatibility name for callers that patch the old setting.  It is
# deliberately not consulted for authorization.
B_UNLOCK_LOCK = os.path.join(HABITAT_ROOT, "b_validation_unlock.json")
MODEL_FREEZE_LOCK = os.path.join(PROGNOSIS_ROOT, "model_freeze_lock.json")

# The W07 split is a frozen A-only artifact with one row per patient per
# repeat/fold/role.  It cannot use the ordinary unique-ID reader because a
# patient_id intentionally occurs in five rows per repeat.
FROZEN_A_SPLIT_RELATIVE_PATH = os.path.join(
    "prognosis_analysis", "output", "outer_splits_A.csv")
FROZEN_A_SPLIT_SHA256 = (
    "24764ee31381621d6a71098a00277743b126a8f00c382afb89d819357ece6502")
FROZEN_A_SPLIT_CANONICAL_SHA256 = FROZEN_A_SPLIT_SHA256
FROZEN_A_SPLIT_COLUMNS = ["patient_id", "repeat", "fold", "role", "seed"]
FROZEN_A_SPLIT_ROLES = ("train", "validation")
FROZEN_A_SPLIT_REPEATS = 10
FROZEN_A_SPLIT_FOLDS = 5
FROZEN_A_SPLIT_SEED_ROOT = 12345

# The P5 technical preflight consumes repeated supervoxel rows: one patient
# may legitimately have several rows, but each (patient, sv_label) pair is a
# unique technical record.  This reader is deliberately separate from the
# ordinary unique-ID reader below.
FROZEN_A_SUPERVOXEL_RELATIVE_PATH = os.path.join(
    "habitat_analysis", "output", "local_global_diagnostic_A_post_slic_fix",
    "supervoxel_mean_A.csv")
FROZEN_A_SUPERVOXEL_COLUMNS = [
    "影像号", "reader", "sv_label", "n_tumor_voxels", "Mean"]
# The producer currently writes these fixed diagnostic metadata columns in
# addition to the five fields consumed by P5.  They are accepted only as this
# complete known source schema; arbitrary extra columns remain invalid.
FROZEN_A_SUPERVOXEL_SOURCE_COLUMNS = [
    "影像号", "reader", "sv_label", "sv_total_voxels",
    "n_tumor_voxels", "physical_volume_mm3", "Mean",
    "slic_requested_scale_mm", "slic_supergrid_voxels_xyz",
    "slic_actual_supergrid_mm_xyz",
]


def resolve_cohort_membership(manifest, scanner):
    """Assign A/B using the single project-wide scanner rule."""
    manifest = manifest.copy()
    scanner = scanner.copy()
    manifest["影像号"] = manifest["影像号"].astype(str).str.strip()
    scanner["影像号"] = scanner["影像号"].astype(str).str.strip()
    if manifest["影像号"].eq("").any() or scanner["影像号"].eq("").any():
        raise AssertionError("manifest/scanner identifiers must be nonempty")
    if manifest["影像号"].duplicated().any() or scanner["影像号"].duplicated().any():
        raise AssertionError("manifest/scanner identifiers must be unique")
    fields = ["影像号", "R1厂商", "R1机型", "R1场强"]
    merged = manifest.merge(scanner[fields], on="影像号", how="left",
                            validate="one_to_one", indicator=True)
    target = merged["排除"].fillna("0").ne("1") if "排除" in merged else pd.Series(True, index=merged.index)
    missing = merged.loc[target & merged["_merge"].ne("both"), "影像号"].tolist()
    if missing:
        raise AssertionError("target cases missing scanner mapping: %s" % missing[:5])
    field = pd.to_numeric(merged["R1场强"], errors="coerce")
    is_a = ((merged["R1厂商"] == "GE MEDICAL SYSTEMS") &
            (merged["R1机型"] == "DISCOVERY MR750") &
            (field.round(1) == 3.0))
    merged["split"] = "B"
    merged.loc[is_a, "split"] = "A"
    return merged.drop(columns=["_merge"])


def add_split(manifest, scanner):
    return resolve_cohort_membership(manifest, scanner)


def require_b_unlock():
    """Require both locks; the legacy B unlock file has no authority."""
    validate_freeze_lock(FREEZE_LOCK)
    return validate_model_freeze_lock(MODEL_FREEZE_LOCK)


def require_a_outcome_unlock():
    """Require the first-stage technical lock before A clinical/outcome reads."""
    return validate_freeze_lock(FREEZE_LOCK)


def _normalize_allowlist(allowed_ids: Optional[Iterable]) -> Optional[Set[str]]:
    if allowed_ids is None:
        return None
    if isinstance(allowed_ids, str):
        values = [allowed_ids]
    else:
        values = list(allowed_ids)
    normalized = {str(value).strip() for value in values}
    if "" in normalized:
        raise ValueError("allowed_ids must not contain an empty identifier")
    return normalized


def _apply_dtype(frame: pd.DataFrame, dtype) -> pd.DataFrame:
    if dtype is None:
        return frame
    if isinstance(dtype, dict):
        for column, target in dtype.items():
            if column in frame.columns:
                frame[column] = frame[column].astype(target)
        return frame
    return frame.astype(dtype)


def _validate_authorized_frame(frame: pd.DataFrame, allowed_ids,
                               id_column: str) -> pd.DataFrame:
    """Validate the output contract at the authorized-read boundary."""
    if not isinstance(frame, pd.DataFrame):
        raise RuntimeError("authorized read must return a pandas DataFrame")
    if id_column not in frame.columns:
        raise RuntimeError("authorized data lacks the identifier column")

    identifiers = frame[id_column]
    if identifiers.isna().any():
        raise RuntimeError("authorized data contains a missing identifier")
    identifiers = identifiers.astype(str).str.strip()
    if identifiers.eq("").any():
        raise RuntimeError("authorized data contains an empty identifier")
    if "读者" in frame.columns:
        uniqueness_key = pd.DataFrame({
            id_column: identifiers,
            "读者": frame["读者"].astype(str).str.strip(),
        }, index=frame.index)
    else:
        uniqueness_key = identifiers
    if uniqueness_key.duplicated().any():
        raise RuntimeError("authorized data contains duplicate identifiers")
    if allowed_ids is not None and not set(identifiers).issubset(allowed_ids):
        raise RuntimeError("authorized data contains an identifier outside the allow-list")

    validated = frame.copy()
    validated[id_column] = identifiers
    return validated


def _absolute_path(path):
    return os.path.normcase(os.path.abspath(os.fspath(path)))


def _sha256_file(path):
    digest = hashlib.sha256()
    with open(path, "rb") as handle:
        for block in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def _canonical_csv_sha256(frame, columns):
    payload = frame[columns].to_csv(index=False)
    payload = payload.replace("\r\n", "\n").replace("\r", "\n")
    return hashlib.sha256(payload.encode("utf-8")).hexdigest()


def _parse_frozen_split_int(raw, column):
    value = "" if raw is None else str(raw)
    if not re.match(r"^(0|[1-9][0-9]*)$", value):
        raise RuntimeError("frozen W07 split has invalid %s" % column)
    return int(value)


def _validate_frozen_split_contract(frame, allowed_ids):
    if list(frame.columns) != FROZEN_A_SPLIT_COLUMNS:
        raise RuntimeError("frozen W07 split columns do not match the contract")
    allowed = _normalize_allowlist(allowed_ids)
    if not allowed:
        raise RuntimeError("frozen W07 split requires a non-empty A allow-list")
    if len(frame) != len(allowed) * FROZEN_A_SPLIT_REPEATS * FROZEN_A_SPLIT_FOLDS:
        raise RuntimeError("frozen W07 split row count does not match the A contract")
    if set(frame["patient_id"]) != allowed:
        raise RuntimeError("frozen W07 split IDs do not match the authorized A IDs")
    if frame.duplicated(subset=["repeat", "fold", "patient_id"]).any():
        raise RuntimeError("frozen W07 split contains duplicate patient/fold rows")

    for repeat in range(1, FROZEN_A_SPLIT_REPEATS + 1):
        repeated = frame[frame["repeat"] == repeat]
        if len(repeated) != len(allowed) * FROZEN_A_SPLIT_FOLDS:
            raise RuntimeError("frozen W07 split repeat coverage mismatch")
        if set(repeated["seed"]) != {
                FROZEN_A_SPLIT_SEED_ROOT + repeat - 1}:
            raise RuntimeError("frozen W07 split seed derivation mismatch")
        validation_counts = repeated[repeated["role"] == "validation"] \
            .groupby("patient_id").size()
        training_counts = repeated[repeated["role"] == "train"] \
            .groupby("patient_id").size()
        if not validation_counts.eq(1).all() or not training_counts.eq(
                FROZEN_A_SPLIT_FOLDS - 1).all():
            raise RuntimeError("frozen W07 split role coverage mismatch")
        for fold in range(1, FROZEN_A_SPLIT_FOLDS + 1):
            group = repeated[repeated["fold"] == fold]
            if len(group) != len(allowed) or set(group["patient_id"]) != allowed:
                raise RuntimeError("frozen W07 split fold coverage mismatch")
            if set(group["role"]) != set(FROZEN_A_SPLIT_ROLES):
                raise RuntimeError("frozen W07 split roles do not match the A contract")


def read_frozen_A_split(path, project_root, allowed_ids=None):
    """Read the immutable A-only W07 split with its repeated-row contract."""
    expected_path = os.path.join(project_root, FROZEN_A_SPLIT_RELATIVE_PATH)
    supplied_path = _absolute_path(path)
    if supplied_path != _absolute_path(expected_path) or \
            _absolute_path(os.path.realpath(path)) != \
            _absolute_path(os.path.realpath(expected_path)):
        raise RuntimeError("frozen W07 split path is not the project-locked A artifact")
    if not os.path.isfile(path):
        raise RuntimeError("frozen W07 split artifact is missing")
    if _sha256_file(path).lower() != FROZEN_A_SPLIT_SHA256:
        raise RuntimeError("frozen W07 split raw hash mismatch")
    allowlist = _normalize_allowlist(allowed_ids)
    if allowlist is None:
        raise RuntimeError("frozen W07 split requires an A allow-list")

    rows = []
    with open(path, "r", newline="", encoding="utf-8-sig") as handle:
        reader = csv.DictReader(handle)
        if reader.fieldnames != FROZEN_A_SPLIT_COLUMNS:
            raise RuntimeError("frozen W07 split columns do not match the contract")
        for row in reader:
            if None in row or any(row.get(column) is None
                                  for column in FROZEN_A_SPLIT_COLUMNS):
                raise RuntimeError("frozen W07 split contains malformed rows")
            patient_id = row["patient_id"]
            if patient_id == "" or patient_id != patient_id.strip():
                raise RuntimeError("frozen W07 split contains invalid patient_id")
            if patient_id not in allowlist:
                raise RuntimeError("frozen W07 split contains an unauthorized patient_id")
            role = row["role"]
            if role not in FROZEN_A_SPLIT_ROLES:
                raise RuntimeError("frozen W07 split contains an invalid role")
            rows.append({
                "patient_id": patient_id,
                "repeat": _parse_frozen_split_int(row["repeat"], "repeat"),
                "fold": _parse_frozen_split_int(row["fold"], "fold"),
                "role": role,
                "seed": _parse_frozen_split_int(row["seed"], "seed"),
            })

    frame = pd.DataFrame(rows, columns=FROZEN_A_SPLIT_COLUMNS)
    _validate_frozen_split_contract(frame, allowlist)
    if _canonical_csv_sha256(frame, FROZEN_A_SPLIT_COLUMNS).lower() != \
            FROZEN_A_SPLIT_CANONICAL_SHA256:
        raise RuntimeError("frozen W07 split canonical hash mismatch")
    return frame


def _parse_frozen_supervoxel_number(raw, column, integer=False, positive=False):
    value = "" if raw is None else str(raw).strip()
    if not value:
        raise RuntimeError("frozen A supervoxel has invalid %s" % column)
    try:
        number = float(value)
    except (TypeError, ValueError):
        raise RuntimeError("frozen A supervoxel has invalid %s" % column)
    if not math.isfinite(number):
        raise RuntimeError("frozen A supervoxel has invalid %s" % column)
    if integer and not number.is_integer():
        raise RuntimeError("frozen A supervoxel has non-integer %s" % column)
    if positive and number <= 0:
        raise RuntimeError("frozen A supervoxel has nonpositive %s" % column)
    return int(number) if integer else number


def read_frozen_A_supervoxels(path, project_root, allowed_ids=None,
                              reader=None, *args, **kwargs):
    """Read the locked A supervoxel source with its repeated-row contract.

    Unlike ordinary technical sources, this artifact is keyed by
    ``(影像号, sv_label)`` rather than by ``影像号`` alone.  The source path,
    header, reader, numeric fields, and key uniqueness are all checked before
    a DataFrame is returned.  Rows outside the authorized A allow-list are
    rejected while streaming and are never returned.
    """
    if reader is not None:
        raise RuntimeError(
            "arbitrary custom readers are not authorized; use the built-in "
            "frozen A supervoxel CSV reader"
        )
    if args or kwargs:
        raise ValueError(
            "frozen A supervoxel reader accepts no custom read options")
    allowlist = _normalize_allowlist(allowed_ids)
    if not allowlist:
        raise RuntimeError(
            "frozen A supervoxel read requires a non-empty A allow-list")

    expected_path = os.path.join(
        os.fspath(project_root), FROZEN_A_SUPERVOXEL_RELATIVE_PATH)
    supplied_path = _absolute_path(path)
    if supplied_path != _absolute_path(expected_path) or \
            _absolute_path(os.path.realpath(path)) != \
            _absolute_path(os.path.realpath(expected_path)):
        raise RuntimeError(
            "frozen A supervoxel path is not the project-locked A artifact")
    if not os.path.isfile(path):
        raise RuntimeError("frozen A supervoxel artifact is missing")
    if os.path.splitext(str(path))[1].lower() != ".csv":
        raise RuntimeError("frozen A supervoxel artifact must be CSV")

    rows = []
    seen_keys = set()
    with open(path, "r", newline="", encoding="utf-8-sig") as handle:
        source = csv.DictReader(handle)
        fieldnames = source.fieldnames
        if fieldnames not in (FROZEN_A_SUPERVOXEL_COLUMNS,
                              FROZEN_A_SUPERVOXEL_SOURCE_COLUMNS):
            raise RuntimeError(
                "frozen A supervoxel columns do not match the contract")
        for row in source:
            if None in row or any(row.get(column) is None
                                  for column in fieldnames):
                raise RuntimeError(
                    "frozen A supervoxel contains malformed rows")

            identifier = str(row["影像号"]).strip()
            if not identifier:
                raise RuntimeError(
                    "frozen A supervoxel contains a blank identifier")
            if identifier not in allowlist:
                raise RuntimeError(
                    "frozen A supervoxel contains an identifier outside the "
                    "allow-list")
            if str(row["reader"]).strip() != "R1":
                raise RuntimeError(
                    "frozen A supervoxel contains a non-R1 reader")

            label = _parse_frozen_supervoxel_number(
                row["sv_label"], "sv_label", integer=True)
            support = _parse_frozen_supervoxel_number(
                row["n_tumor_voxels"], "n_tumor_voxels",
                integer=True, positive=True)
            mean = _parse_frozen_supervoxel_number(row["Mean"], "Mean")
            key = (identifier, label)
            if key in seen_keys:
                raise RuntimeError(
                    "frozen A supervoxel contains duplicate patient_id+sv_label")
            seen_keys.add(key)
            rows.append({
                "影像号": identifier,
                "reader": "R1",
                "sv_label": label,
                "n_tumor_voxels": support,
                "Mean": mean,
            })

    return pd.DataFrame(rows, columns=FROZEN_A_SUPERVOXEL_COLUMNS)


def _stream_csv(path, allowed_ids: Set[str], id_column: str, *args, **kwargs):
    encoding = kwargs.pop("encoding", "utf-8-sig")
    dtype = kwargs.pop("dtype", None)
    usecols = kwargs.pop("usecols", None)
    delimiter = kwargs.pop("delimiter", ",")
    if args or kwargs:
        raise ValueError("unsupported options for authorized CSV reader: %s" %
                         sorted(kwargs))
    rows = []
    with open(path, "r", newline="", encoding=encoding) as handle:
        reader = csv.DictReader(handle, delimiter=delimiter)
        if not reader.fieldnames or id_column not in reader.fieldnames:
            raise ValueError("authorized CSV source lacks the identifier column")
        columns = list(reader.fieldnames)
        if usecols is not None:
            columns = [column for column in columns if column in set(usecols)]
            if id_column not in columns:
                columns.insert(0, id_column)
        for row in reader:
            identifier = str(row.get(id_column, "")).strip()
            if identifier in allowed_ids:
                rows.append({column: row.get(column) for column in columns})
    return _apply_dtype(pd.DataFrame(rows, columns=columns), dtype)


def _stream_excel(path, allowed_ids: Set[str], id_column: str, *args, **kwargs):
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("openpyxl is required for authorized XLSX reads") from exc
    if WorkSheetParser is None or coordinate_to_tuple is None:
        raise RuntimeError("openpyxl worksheet parser is required for authorized XLSX reads")
    if args:
        raise ValueError("positional options are unsupported for authorized XLSX reader")
    dtype = kwargs.pop("dtype", None)
    sheet_name = kwargs.pop("sheet_name", 0)
    usecols = kwargs.pop("usecols", None)
    if kwargs:
        raise ValueError("unsupported options for authorized XLSX reader: %s" %
                         sorted(kwargs))
    if isinstance(sheet_name, (list, tuple)):
        raise ValueError("authorized XLSX reader accepts one sheet only")
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        if isinstance(sheet_name, int):
            worksheet = workbook.worksheets[sheet_name]
        else:
            worksheet = workbook[sheet_name]
        class _AuthorizedWorksheetParser(WorkSheetParser):
            """Parse only selected cells from selected, already-authorized rows."""

            def __init__(self, *parser_args, selected_rows=None,
                         selected_columns=None, **parser_kwargs):
                super().__init__(*parser_args, **parser_kwargs)
                self.selected_rows = selected_rows
                self.selected_columns = (None if selected_columns is None
                                         else set(selected_columns))

            def parse_row(self, row):
                attrs = dict(row.attrib)
                if "r" in attrs:
                    try:
                        self.row_counter = int(attrs["r"])
                    except ValueError:
                        value = float(attrs["r"])
                        if not value.is_integer():
                            raise
                        self.row_counter = int(value)
                else:
                    self.row_counter += 1
                self.col_counter = 0
                if (self.selected_rows is not None and
                        self.row_counter not in self.selected_rows):
                    return self.row_counter, []

                cells = []
                for element in row:
                    coordinate = element.get("r")
                    if coordinate:
                        _, column = coordinate_to_tuple(coordinate)
                        self.col_counter = column
                    else:
                        self.col_counter += 1
                        column = self.col_counter
                    if (self.selected_columns is not None and
                            column not in self.selected_columns):
                        continue
                    # parse_cell increments col_counter when a cell has no
                    # explicit coordinate; restore the pre-cell position so
                    # sparse coordinate-less rows remain correctly indexed.
                    if not coordinate:
                        self.col_counter = column - 1
                    cells.append(self.parse_cell(element))
                return self.row_counter, cells

        def filtered_rows(selected_rows, selected_columns):
            source = worksheet._get_source()
            parser = _AuthorizedWorksheetParser(
                source,
                worksheet._shared_strings,
                data_only=worksheet.parent.data_only,
                epoch=worksheet.parent.epoch,
                date_formats=worksheet.parent._date_formats,
                selected_rows=selected_rows,
                selected_columns=selected_columns,
            )
            try:
                for row_number, cells in parser.parse():
                    if selected_rows is None or row_number in selected_rows:
                        yield row_number, {
                            cell["column"]: cell["value"] for cell in cells
                        }
            finally:
                source.close()

        header_rows = filtered_rows({1}, None)
        try:
            _, header_values = next(header_rows)
        except StopIteration:
            raise ValueError("authorized XLSX source is empty")
        finally:
            header_rows.close()
        header_width = worksheet.max_column or (max(header_values) if header_values else 0)
        header = [str(header_values.get(number, "")).strip()
                  if header_values.get(number) is not None else ""
                  for number in range(1, header_width + 1)]
        if id_column not in header:
            raise ValueError("authorized XLSX source lacks the identifier column")
        if header.count(id_column) != 1:
            raise ValueError("authorized XLSX source must contain exactly one identifier column")
        if usecols is None:
            column_numbers = list(range(1, len(header) + 1))
        else:
            requested = {usecols} if isinstance(usecols, str) else set(usecols)
            column_numbers = [number for number, column in enumerate(header, 1)
                              if column in requested]
            id_number = header.index(id_column) + 1
            if id_number not in column_numbers:
                column_numbers.insert(0, id_number)
        columns = [header[number - 1] for number in column_numbers]

        # First pass touches only the identifier column.  No non-ID cell from
        # any row is parsed while the A allow-list is being resolved.
        id_number = header.index(id_column) + 1
        selected_rows = []
        for row_number, cell_values in filtered_rows(None, {id_number}):
            identifier = str(cell_values.get(id_number, "")).strip()
            if identifier in allowed_ids:
                selected_rows.append(row_number)

        # Second pass touches only requested columns and only rows admitted by
        # the first pass.  Non-A rows never become application-level records.
        output = []
        selected_row_set = set(selected_rows)
        for _, cell_values in filtered_rows(selected_row_set, set(column_numbers)):
            output.append([cell_values.get(number) for number in column_numbers])
        return _apply_dtype(pd.DataFrame(output, columns=columns), dtype)
    finally:
        workbook.close()


def _authorized_read(path, reader, allowed_ids, id_column, allow_full,
                     *args, **kwargs):
    allowlist = _normalize_allowlist(allowed_ids)
    if reader is not None:
        # A callable may materialize disallowed rows before this function can
        # inspect or filter its return value.  Reject it before execution.
        raise RuntimeError(
            "arbitrary custom readers are not authorized; use the built-in "
            "CSV/XLSX streaming reader"
        )
    if allowlist is None:
        if not allow_full:
            raise RuntimeError("authorized read requires an identifier allow-list")
        suffix = os.path.splitext(str(path))[1].lower()
        if suffix == ".csv":
            result = pd.read_csv(path, *args, **kwargs)
            return _validate_authorized_frame(result, None, id_column)
        if suffix in (".xlsx", ".xlsm"):
            result = pd.read_excel(path, *args, **kwargs)
            return _validate_authorized_frame(result, None, id_column)
        raise ValueError("unsupported declared technical format: %s" % suffix)
    suffix = os.path.splitext(str(path))[1].lower()
    if suffix in (".csv", ".tsv"):
        if suffix == ".tsv" and "delimiter" not in kwargs:
            kwargs["delimiter"] = "\t"
        result = _stream_csv(path, allowlist, id_column, *args, **kwargs)
        return _validate_authorized_frame(result, allowlist, id_column)
    if suffix in (".xlsx", ".xlsm"):
        result = _stream_excel(path, allowlist, id_column, *args, **kwargs)
        return _validate_authorized_frame(result, allowlist, id_column)
    raise ValueError("unsupported authorized data format: %s" % suffix)


def read_technical_A(path, reader=None, allowed_ids=None, id_column="影像号",
                     allow_full=False, *args, **kwargs):
    """Read outcome-blind technical A data without a lock.

    ``allowed_ids`` is required for a raw source that may contain A and B.
    A declared A-only technical artifact may be read with ``allow_full=True``.
    The legacy ``reader`` argument is retained for call compatibility but is
    rejected; only the built-in source-level streaming reader is authorized.
    """
    return _authorized_read(path, reader, allowed_ids, id_column, allow_full,
                            *args, **kwargs)


def read_A_outcomes(path, reader=None, allowed_ids=None, id_column="影像号",
                    *args, **kwargs):
    """Validate the first lock, then read only authorized A clinical/outcome rows."""
    require_a_outcome_unlock()
    return _authorized_read(path, reader, allowed_ids, id_column, False,
                            *args, **kwargs)


def read_B_validation(path, reader=None, allowed_ids=None, id_column="影像号",
                      *args, **kwargs):
    """Validate the model freeze, then read only authorized B validation rows."""
    require_b_unlock()
    return _authorized_read(path, reader, allowed_ids, id_column, False,
                            *args, **kwargs)


# Compatibility aliases are intentionally non-authorizing; callers should use
# the explicit three-reader API above.
read_technical_data = read_technical_A


def read_a_outcome(path, reader, *args, **kwargs):
    return read_A_outcomes(path, reader=reader, *args, **kwargs)


def read_b_data(path, reader, *args, **kwargs):
    return read_B_validation(path, reader=reader, *args, **kwargs)


def read_b_csv(path, *args, **kwargs):
    reader = kwargs.pop("reader", None)
    return read_B_validation(path, reader=reader, *args, **kwargs)


def read_b_excel(path, *args, **kwargs):
    reader = kwargs.pop("reader", None)
    return read_B_validation(path, reader=reader, *args, **kwargs)


def select_split(frame, split):
    if split not in ("A", "B", "all"):
        raise ValueError("split must be A, B, or all")
    if split in ("B", "all"):
        require_b_unlock()
    return frame.copy() if split == "all" else frame[frame["split"] == split].copy()
